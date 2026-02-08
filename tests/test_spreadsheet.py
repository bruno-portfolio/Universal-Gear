"""Tests for the spreadsheet template generator and JSON importer."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from universal_gear.cli.spreadsheet import (
    SHEET_NAMES,
    generate_template,
    read_sheet_as_json,
)

EXPECTED_SHEETS = 7


@pytest.fixture()
def template_path(tmp_path: Path) -> Path:
    return generate_template(tmp_path / "test_template.xlsx")


@pytest.mark.offline()
class TestGenerateTemplate:
    def test_creates_file(self, tmp_path: Path):
        path = generate_template(tmp_path / "out.xlsx")
        assert path.exists()
        assert path.suffix == ".xlsx"

    def test_has_all_seven_sheets(self, template_path: Path):
        wb = load_workbook(str(template_path))
        assert len(wb.sheetnames) == EXPECTED_SHEETS

    def test_sheet_names_match(self, template_path: Path):
        wb = load_workbook(str(template_path))
        assert tuple(wb.sheetnames) == SHEET_NAMES

    def test_each_sheet_has_instruction_row(self, template_path: Path):
        wb = load_workbook(str(template_path))
        for name in SHEET_NAMES:
            ws = wb[name]
            first_cell = ws.cell(row=1, column=1).value
            assert first_cell is not None, f"Sheet '{name}' missing instruction"
            assert len(str(first_cell)) > 10, f"Sheet '{name}' instruction too short"

    def test_observe_has_example_data(self, template_path: Path):
        wb = load_workbook(str(template_path))
        ws = wb[SHEET_NAMES[0]]
        values = [ws.cell(row=r, column=5).value for r in range(3, 6)]
        assert 32.50 in values

    def test_simulate_has_three_scenarios(self, template_path: Path):
        wb = load_workbook(str(template_path))
        ws = wb[SHEET_NAMES[3]]
        scenario_names = []
        for row in range(3, 10):
            val = ws.cell(row=row, column=1).value
            if val and isinstance(val, str) and val in ("Otimista", "Base", "Pessimista"):
                scenario_names.append(val)
        assert len(scenario_names) == 3

    def test_english_language(self, tmp_path: Path):
        path = generate_template(tmp_path / "en.xlsx", lang="en")
        wb = load_workbook(str(path))
        ws = wb[SHEET_NAMES[0]]
        instruction = ws.cell(row=1, column=1).value
        assert "OBSERVE" in instruction

    def test_returns_path_object(self, tmp_path: Path):
        path = generate_template(tmp_path / "out.xlsx")
        assert isinstance(path, Path)


@pytest.mark.offline()
class TestReadSheetAsJson:
    def test_roundtrip_reads_observations(self, template_path: Path):
        data = read_sheet_as_json(template_path)
        assert "observations" in data
        assert len(data["observations"]) >= 3

    def test_roundtrip_reads_scenarios(self, template_path: Path):
        data = read_sheet_as_json(template_path)
        assert "scenarios" in data
        assert len(data["scenarios"]) >= 2

    def test_roundtrip_reads_decisions(self, template_path: Path):
        data = read_sheet_as_json(template_path)
        assert "decisions" in data
        assert len(data["decisions"]) >= 1

    def test_roundtrip_reads_feedback(self, template_path: Path):
        data = read_sheet_as_json(template_path)
        assert "feedback" in data
        assert len(data["feedback"]) >= 1

    def test_all_sections_present(self, template_path: Path):
        data = read_sheet_as_json(template_path)
        expected_keys = {
            "observations", "compressions", "hypotheses",
            "scenarios", "decisions", "feedback", "dashboard",
        }
        assert set(data.keys()) == expected_keys

    def test_observation_has_expected_fields(self, template_path: Path):
        data = read_sheet_as_json(template_path)
        obs = data["observations"][0]
        assert "Fonte" in obs or "Data" in obs
