"""Export helpers for pipeline results (JSON / CSV / XLSX)."""

from __future__ import annotations

import csv
import io
import json
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from pathlib import Path

    from universal_gear.core.pipeline import PipelineResult


def _serialize_stage(obj: object | None) -> dict[str, Any] | None:
    """Serialize a Pydantic contract object to a JSON-safe dict."""
    if obj is None:
        return None
    if hasattr(obj, "model_dump"):
        return obj.model_dump(mode="json")  # type: ignore[union-attr]
    return None


def _build_payload(result: PipelineResult) -> dict[str, Any]:
    """Build the full JSON payload from a PipelineResult."""
    stages_data: dict[str, Any] = {
        "observation": _serialize_stage(result.collection),
        "compression": _serialize_stage(result.compression),
        "hypothesis": _serialize_stage(result.hypothesis),
        "simulation": _serialize_stage(result.simulation),
        "decision": _serialize_stage(result.decision),
        "feedback": _serialize_stage(result.feedback),
    }

    metrics_data: dict[str, Any] = {
        "total_duration": result.metrics.total_duration,
        "all_success": result.metrics.all_success,
        "stages": [
            {
                "stage": s.stage,
                "duration_seconds": s.duration_seconds,
                "success": s.success,
                "error": s.error,
            }
            for s in result.metrics.stages
        ],
    }

    return {
        "success": result.success,
        "error": result.error,
        "stages": stages_data,
        "metrics": metrics_data,
    }


def export_json(result: PipelineResult) -> str:
    """Serialize the full PipelineResult to a JSON string."""
    payload = _build_payload(result)
    return json.dumps(payload, indent=2, ensure_ascii=False)


def export_csv(result: PipelineResult) -> str:
    """Serialize the PipelineResult as a CSV summary table."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["stage", "status", "detail", "duration_seconds"])

    for stage_metric in result.metrics.stages:
        status = "OK" if stage_metric.success else "FAIL"
        detail = _stage_detail_plain(result, stage_metric.stage)
        writer.writerow(
            [
                stage_metric.stage,
                status,
                detail,
                f"{stage_metric.duration_seconds:.3f}",
            ]
        )

    writer.writerow(
        [
            "TOTAL",
            "SUCCESS" if result.success else "FAILED",
            result.error or "",
            f"{result.metrics.total_duration:.3f}",
        ]
    )

    return buf.getvalue()


def _stage_detail_plain(  # noqa: PLR0911
    result: PipelineResult,
    stage: str,
) -> str:
    """Plain-text stage detail (no Rich markup)."""
    match stage:
        case "observation":
            if result.collection:
                n = len(result.collection.events)
                rel = result.collection.quality_report.reliability_score
                return f"{n} events | reliability: {rel:.2f}"
        case "compression":
            if result.compression:
                n = len(result.compression.states)
                gran = (
                    result.compression.states[0].granularity.value
                    if result.compression.states
                    else "?"
                )
                return f"{n} states | {gran}"
        case "hypothesis":
            if result.hypothesis:
                n = len(result.hypothesis.hypotheses)
                return f"{n} hypotheses"
        case "simulation":
            if result.simulation:
                n = len(result.simulation.scenarios)
                has_bl = "baseline + " if result.simulation.baseline else ""
                return f"{has_bl}{n} scenarios"
        case "decision":
            if result.decision:
                n = len(result.decision.decisions)
                types = {d.decision_type.value for d in result.decision.decisions}
                return f"{n} decisions | {', '.join(types)}"
        case "feedback":
            if result.feedback:
                n = len(result.feedback.scorecards)
                return f"{n} scorecards"
    return ""


_ACTIONABLE_TYPES = {"recommendation", "trigger", "alert"}


def export_xlsx(result: PipelineResult, output_path: Path) -> Path:
    """Export PipelineResult to a formatted xlsx workbook with 7 tabs."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    _xlsx_observe(wb, result)
    _xlsx_compress(wb, result)
    _xlsx_hypothesis(wb, result)
    _xlsx_simulate(wb, result)
    _xlsx_decide(wb, result)
    _xlsx_feedback(wb, result)
    _xlsx_dashboard(wb, result)

    wb.save(str(output_path))
    return output_path


def _xlsx_setup(
    wb: Any, sheet_idx: int, instruction: str,
    headers: list[str], widths: list[int],
) -> tuple[Any, int]:
    from universal_gear.cli.spreadsheet import (
        SHEET_NAMES,
        _add_headers,
        _add_instruction,
        _set_col_widths,
    )

    ws = wb.create_sheet(SHEET_NAMES[sheet_idx])
    _set_col_widths(ws, widths)
    row = _add_instruction(ws, 1, instruction, len(headers))
    row = _add_headers(ws, row, headers)
    return ws, row


def _xlsx_data_rows(ws: Any, start: int, rows: list[list[Any]]) -> int:
    for row_values in rows:
        for col_idx, val in enumerate(row_values, start=1):
            ws.cell(row=start, column=col_idx, value=val)
        start += 1
    return start


def _xlsx_observe(wb: Any, result: PipelineResult) -> None:
    from universal_gear.cli.spreadsheet import (
        COL_WIDTH_MEDIUM,
        COL_WIDTH_NARROW,
        COL_WIDTH_WIDE,
    )

    headers = [
        "Data", "Fonte", "Tipo", "Descricao",
        "Valor", "Unidade", "Confiavel",
    ]
    widths = [
        COL_WIDTH_NARROW, COL_WIDTH_MEDIUM, COL_WIDTH_NARROW,
        COL_WIDTH_WIDE, COL_WIDTH_NARROW, COL_WIDTH_NARROW,
        COL_WIDTH_NARROW,
    ]
    ws, row = _xlsx_setup(
        wb, 0, "OBSERVAR: Dados brutos coletados pelo pipeline.",
        headers, widths,
    )
    if not result.collection:
        return

    rows: list[list[Any]] = []
    for evt in result.collection.events:
        ts = evt.timestamp.strftime("%Y-%m-%d %H:%M")
        rows.append([
            ts, evt.source.source_id, evt.source.source_type.value,
            evt.data.get("description", evt.data.get("metric", "")),
            evt.data.get("value", ""), evt.data.get("unit", ""),
            evt.source.reliability.value,
        ])
    _xlsx_data_rows(ws, row, rows)


def _xlsx_compress(wb: Any, result: PipelineResult) -> None:
    from universal_gear.cli.spreadsheet import COL_WIDTH_MEDIUM, COL_WIDTH_NARROW

    headers = [
        "Periodo", "Metrica", "Valor",
        "Unidade", "Confianca", "Confiabilidade",
    ]
    widths = [
        COL_WIDTH_MEDIUM, COL_WIDTH_MEDIUM, COL_WIDTH_NARROW,
        COL_WIDTH_NARROW, COL_WIDTH_NARROW, COL_WIDTH_NARROW,
    ]
    ws, row = _xlsx_setup(
        wb, 1, "COMPRIMIR: Estados de mercado normalizados.",
        headers, widths,
    )
    if not result.compression:
        return

    rows: list[list[Any]] = []
    for state in result.compression.states:
        start = state.period_start.strftime("%Y-%m-%d")
        end = state.period_end.strftime("%Y-%m-%d")
        period = f"{start} - {end}"
        for sig in state.signals:
            rows.append([
                period, sig.name, sig.value, sig.unit,
                f"{sig.confidence:.0%}",
                f"{state.source_reliability:.0%}",
            ])
    _xlsx_data_rows(ws, row, rows)


def _xlsx_hypothesis(wb: Any, result: PipelineResult) -> None:
    from universal_gear.cli.spreadsheet import COL_WIDTH_NARROW, COL_WIDTH_WIDE

    headers = [
        "Hipotese", "Justificativa",
        "Confianca %", "Status", "Valida ate",
    ]
    widths = [
        COL_WIDTH_WIDE, COL_WIDTH_WIDE,
        COL_WIDTH_NARROW, COL_WIDTH_NARROW, COL_WIDTH_NARROW,
    ]
    ws, row = _xlsx_setup(
        wb, 2, "HIPOTESE: Hipoteses testaveis geradas pelo pipeline.",
        headers, widths,
    )
    if not result.hypothesis:
        return

    rows: list[list[Any]] = []
    for hyp in result.hypothesis.hypotheses:
        rows.append([
            hyp.statement, hyp.rationale, f"{hyp.confidence:.0%}",
            hyp.status.value, hyp.valid_until.strftime("%Y-%m-%d"),
        ])
    _xlsx_data_rows(ws, row, rows)


def _xlsx_simulate(wb: Any, result: PipelineResult) -> None:
    from universal_gear.cli.spreadsheet import (
        COL_WIDTH_MEDIUM,
        COL_WIDTH_NARROW,
        COL_WIDTH_WIDE,
    )

    headers = [
        "Cenario", "Descricao", "Premissas",
        "Resultado", "Probabilidade %", "Risco",
    ]
    widths = [
        COL_WIDTH_MEDIUM, COL_WIDTH_WIDE, COL_WIDTH_WIDE,
        COL_WIDTH_WIDE, COL_WIDTH_NARROW, COL_WIDTH_NARROW,
    ]
    ws, row = _xlsx_setup(
        wb, 3, "SIMULAR: Cenarios condicionais projetados.",
        headers, widths,
    )
    if not result.simulation:
        return

    scenarios = list(result.simulation.scenarios)
    if result.simulation.baseline:
        scenarios.insert(0, result.simulation.baseline)

    rows: list[list[Any]] = []
    for sc in scenarios:
        assumptions = "; ".join(
            f"{a.variable}={a.assumed_value}" for a in sc.assumptions
        )
        outcomes = ", ".join(
            f"{k}: {v}" for k, v in sc.projected_outcome.items()
        )
        rows.append([
            sc.name, sc.description, assumptions, outcomes,
            f"{sc.probability:.0%}", sc.risk_level.value,
        ])
    _xlsx_data_rows(ws, row, rows)


def _xlsx_decide(wb: Any, result: PipelineResult) -> None:
    from universal_gear.cli.spreadsheet import COL_WIDTH_NARROW, COL_WIDTH_WIDE

    headers = [
        "Decisao", "Tipo", "Recomendacao", "Confianca %",
        "Risco", "Custo Falso Positivo", "Custo Falso Negativo",
    ]
    widths = [
        COL_WIDTH_WIDE, COL_WIDTH_NARROW, COL_WIDTH_WIDE,
        COL_WIDTH_NARROW, COL_WIDTH_NARROW,
        COL_WIDTH_WIDE, COL_WIDTH_WIDE,
    ]
    ws, row = _xlsx_setup(
        wb, 4, "DECIDIR: Decisoes geradas pelo pipeline.",
        headers, widths,
    )
    if not result.decision:
        return

    rows: list[list[Any]] = []
    for dec in result.decision.decisions:
        rows.append([
            dec.title, dec.decision_type.value,
            dec.recommendation, f"{dec.confidence:.0%}",
            dec.risk_level.value,
            dec.cost_of_error.false_positive,
            dec.cost_of_error.false_negative,
        ])
    _xlsx_data_rows(ws, row, rows)


def _xlsx_feedback(wb: Any, result: PipelineResult) -> None:
    from universal_gear.cli.spreadsheet import (
        COL_WIDTH_MEDIUM,
        COL_WIDTH_NARROW,
        COL_WIDTH_WIDE,
    )

    headers = [
        "Decisao ID", "Metrica", "Previsto", "Real",
        "Erro %", "Acertou?", "Resultado",
    ]
    widths = [
        COL_WIDTH_WIDE, COL_WIDTH_MEDIUM, COL_WIDTH_NARROW,
        COL_WIDTH_NARROW, COL_WIDTH_NARROW, COL_WIDTH_NARROW,
        COL_WIDTH_WIDE,
    ]
    ws, row = _xlsx_setup(
        wb, 5, "FEEDBACK: Comparacao entre previsoes e realidade.",
        headers, widths,
    )
    if not result.feedback:
        return

    rows: list[list[Any]] = []
    for sc in result.feedback.scorecards:
        for pvr in sc.predictions_vs_reality:
            hit = "Sim" if pvr.within_confidence else "Nao"
            rows.append([
                str(sc.decision_id), pvr.metric,
                pvr.predicted, pvr.actual,
                f"{pvr.error_pct:.1f}%", hit,
                sc.decision_outcome,
            ])
    _xlsx_data_rows(ws, row, rows)


def _xlsx_dashboard(wb: Any, result: PipelineResult) -> None:
    from universal_gear.cli.spreadsheet import (
        COL_WIDTH_MEDIUM,
        COL_WIDTH_WIDE,
        _styles,
    )

    headers = ["Metrica", "Valor"]
    widths = [COL_WIDTH_WIDE, COL_WIDTH_MEDIUM]
    ws, row = _xlsx_setup(
        wb, 6, "DASHBOARD: Metricas consolidadas do pipeline.",
        headers, widths,
    )

    total_decisions = 0
    actionable = 0
    if result.decision:
        total_decisions = len(result.decision.decisions)
        actionable = sum(
            1 for d in result.decision.decisions
            if d.decision_type.value in _ACTIONABLE_TYPES
        )

    hit_rate: str | int = "N/A"
    mean_error: str | int = "N/A"
    degradations = 0
    if result.feedback and result.feedback.scorecards:
        all_pvr = [
            pvr
            for sc in result.feedback.scorecards
            for pvr in sc.predictions_vs_reality
        ]
        if all_pvr:
            hits = sum(1 for p in all_pvr if p.within_confidence)
            hit_rate = f"{hits / len(all_pvr):.0%}"
            total_err = sum(p.error_pct for p in all_pvr)
            mean_error = f"{total_err / len(all_pvr):.1f}%"
        degradations = sum(
            len(sc.source_degradations)
            for sc in result.feedback.scorecards
        )

    status = (
        "SUCCESS" if result.success
        else f"FAILED: {result.error or 'unknown'}"
    )

    s = _styles()
    metrics_rows: list[tuple[str, Any]] = [
        ("Total de decisoes", total_decisions),
        ("Decisoes acionaveis", actionable),
        ("Hit rate", hit_rate),
        ("Erro medio", mean_error),
        ("Degradacoes de fonte", degradations),
        ("Duracao total", f"{result.metrics.total_duration:.1f}s"),
        ("Status", status),
    ]
    for label, val in metrics_rows:
        cell_label = ws.cell(row=row, column=1, value=label)
        cell_label.font = s["bold"]
        ws.cell(row=row, column=2, value=val)
        row += 1
