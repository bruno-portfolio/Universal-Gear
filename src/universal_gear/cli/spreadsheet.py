"""Spreadsheet template generator and JSON importer for Universal Gear."""

from __future__ import annotations

from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from pathlib import Path

HEADER_FILL_HEX = "1F4E79"
HEADER_FONT_COLOR = "FFFFFF"
INSTRUCTION_FILL_HEX = "FFF2CC"
INPUT_FILL_HEX = "E2EFDA"
PROTECTED_FILL_HEX = "D9E2F3"
EXAMPLE_FILL_HEX = "FCE4D6"

SHEET_NAMES = (
    "OBSERVAR",
    "COMPRIMIR",
    "HIPOTESE",
    "SIMULAR",
    "DECIDIR",
    "FEEDBACK",
    "DASHBOARD",
)

COL_WIDTH_NARROW = 14
COL_WIDTH_MEDIUM = 22
COL_WIDTH_WIDE = 40
MIN_HEADER_COLS = 2


def generate_template(output_path: Path, *, lang: str = "pt") -> Path:
    """Generate the decision-framework xlsx template and return the file path."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    _build_observe(wb, lang=lang)
    _build_compress(wb, lang=lang)
    _build_hypothesize(wb, lang=lang)
    _build_simulate(wb, lang=lang)
    _build_decide(wb, lang=lang)
    _build_feedback(wb, lang=lang)
    _build_dashboard(wb, lang=lang)

    wb.save(str(output_path))
    return output_path


def _styles():
    from openpyxl.styles import Alignment, Font, PatternFill

    return {
        "header_fill": PatternFill("solid", fgColor=HEADER_FILL_HEX),
        "header_font": Font(bold=True, color=HEADER_FONT_COLOR, size=11),
        "instruction_fill": PatternFill("solid", fgColor=INSTRUCTION_FILL_HEX),
        "input_fill": PatternFill("solid", fgColor=INPUT_FILL_HEX),
        "protected_fill": PatternFill("solid", fgColor=PROTECTED_FILL_HEX),
        "example_fill": PatternFill("solid", fgColor=EXAMPLE_FILL_HEX),
        "bold": Font(bold=True, size=11),
        "wrap": Alignment(wrap_text=True, vertical="top"),
    }


def _add_instruction(ws: Any, row: int, text: str, cols: int) -> int:
    """Write an instruction block across merged cells. Returns next available row."""
    s = _styles()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = s["instruction_fill"]
    cell.alignment = s["wrap"]
    cell.font = s["bold"]
    ws.row_dimensions[row].height = 45
    return row + 1


def _add_headers(ws: Any, row: int, headers: list[str]) -> int:
    """Write styled header row. Returns next available row."""
    s = _styles()
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = s["header_fill"]
        cell.font = s["header_font"]
        cell.alignment = s["wrap"]
    return row + 1


def _set_col_widths(ws: Any, widths: list[int]) -> None:
    from openpyxl.utils import get_column_letter

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _add_example_row(ws: Any, row: int, values: list[Any]) -> int:
    s = _styles()
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.fill = s["example_fill"]
        cell.alignment = s["wrap"]
    return row + 1


def _build_observe(wb: Any, *, lang: str) -> None:
    ws = wb.create_sheet(SHEET_NAMES[0])

    instruction = (
        "OBSERVAR: Registre aqui os dados brutos que voce coletou. "
        "Pode ser preco, quantidade, noticia, qualquer informacao relevante. "
        "Uma linha por observacao. Preencha as celulas verdes."
        if lang == "pt"
        else "OBSERVE: Record raw data here. One row per observation. Fill the green cells."
    )

    headers = ["Data", "Fonte", "Tipo", "Descricao", "Valor", "Unidade", "Confiavel?"]
    widths = [COL_WIDTH_NARROW, COL_WIDTH_MEDIUM, COL_WIDTH_NARROW, COL_WIDTH_WIDE,
              COL_WIDTH_NARROW, COL_WIDTH_NARROW, COL_WIDTH_NARROW]

    _set_col_widths(ws, widths)
    row = _add_instruction(ws, 1, instruction, len(headers))
    row = _add_headers(ws, row, headers)

    examples = [
        ["2024-11-01", "Fornecedor A", "preco", "Cafe arabica 1kg", 32.50, "BRL/kg", "Sim"],
        ["2024-11-05", "Mercado local", "preco", "Cafe arabica 1kg", 34.00, "BRL/kg", "Sim"],
        ["2024-11-10", "Site importador", "preco", "Cafe arabica 1kg", 31.80, "BRL/kg", "Medio"],
    ]
    for ex in examples:
        row = _add_example_row(ws, row, ex)

    s = _styles()
    for r in range(row, row + 20):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = s["input_fill"]


def _build_compress(wb: Any, *, lang: str) -> None:
    ws = wb.create_sheet(SHEET_NAMES[1])

    instruction = (
        "COMPRIMIR: Resuma seus dados. Qual a media? A tendencia esta subindo ou descendo? "
        "Quanto variou? Preencha manualmente ou use formulas. Celulas azuis sao calculadas."
        if lang == "pt"
        else "COMPRESS: Summarize your data. What is the average? Is the trend up or down? "
             "Fill manually or use formulas. Blue cells are calculated."
    )

    headers = ["Periodo", "Metrica", "Media", "Minimo", "Maximo", "Variacao %", "Tendencia"]
    widths = [COL_WIDTH_MEDIUM, COL_WIDTH_MEDIUM, COL_WIDTH_NARROW, COL_WIDTH_NARROW,
              COL_WIDTH_NARROW, COL_WIDTH_NARROW, COL_WIDTH_NARROW]

    _set_col_widths(ws, widths)
    row = _add_instruction(ws, 1, instruction, len(headers))
    row = _add_headers(ws, row, headers)

    examples = [
        ["Nov 2024", "Preco cafe/kg", 32.77, 31.80, 34.00, "6.9%", "Alta"],
    ]
    for ex in examples:
        row = _add_example_row(ws, row, ex)

    s = _styles()
    for r in range(row, row + 10):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = s["input_fill"]


def _build_hypothesize(wb: Any, *, lang: str) -> None:
    ws = wb.create_sheet(SHEET_NAMES[2])

    instruction = (
        "HIPOTESE: O que voce acha que esta acontecendo? Escreva sua hipotese, "
        "o que confirmaria (validacao) e o que provaria que esta errada (falsificacao). "
        "Toda hipotese boa pode ser testada."
        if lang == "pt"
        else "HYPOTHESIZE: What do you think is happening? Write your hypothesis, "
             "what would confirm it (validation) and disprove it (falsification)."
    )

    headers = ["Hipotese", "Justificativa", "Confianca %",
               "Criterio de Validacao", "Criterio de Falsificacao", "Valida ate"]
    widths = [COL_WIDTH_WIDE, COL_WIDTH_WIDE, COL_WIDTH_NARROW,
              COL_WIDTH_WIDE, COL_WIDTH_WIDE, COL_WIDTH_NARROW]

    _set_col_widths(ws, widths)
    row = _add_instruction(ws, 1, instruction, len(headers))
    row = _add_headers(ws, row, headers)

    examples = [
        [
            "Preco do cafe vai subir 10% no proximo mes",
            "Tendencia de alta nos ultimos 3 meses + seca no Cerrado",
            "70%",
            "Preco medio > 35 BRL/kg em Dez 2024",
            "Preco cai abaixo de 30 BRL/kg",
            "2024-12-31",
        ],
    ]
    for ex in examples:
        row = _add_example_row(ws, row, ex)

    s = _styles()
    for r in range(row, row + 10):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = s["input_fill"]


def _build_simulate(wb: Any, *, lang: str) -> None:
    ws = wb.create_sheet(SHEET_NAMES[3])

    instruction = (
        "SIMULAR: Crie pelo menos 2 cenarios (otimista e pessimista). "
        "Para cada um: quais premissas? Qual o resultado esperado? Qual a probabilidade? "
        "Qual o risco?"
        if lang == "pt"
        else "SIMULATE: Create at least 2 scenarios (optimistic and pessimistic). "
             "For each: assumptions, expected outcome, probability, risk level."
    )

    headers = ["Cenario", "Descricao", "Premissas", "Resultado Esperado",
               "Probabilidade %", "Risco"]
    widths = [COL_WIDTH_MEDIUM, COL_WIDTH_WIDE, COL_WIDTH_WIDE, COL_WIDTH_WIDE,
              COL_WIDTH_NARROW, COL_WIDTH_NARROW]

    _set_col_widths(ws, widths)
    row = _add_instruction(ws, 1, instruction, len(headers))
    row = _add_headers(ws, row, headers)

    examples = [
        [
            "Otimista",
            "Preco sobe como esperado",
            "Seca continua; demanda global alta",
            "Cafe a 36 BRL/kg em Dez",
            "40%",
            "Baixo",
        ],
        [
            "Base",
            "Preco estavel com leve alta",
            "Chuva parcial; demanda estavel",
            "Cafe a 33.50 BRL/kg em Dez",
            "40%",
            "Medio",
        ],
        [
            "Pessimista",
            "Preco cai por safra recorde",
            "Chuvas abundantes; oferta excedente",
            "Cafe a 28 BRL/kg em Dez",
            "20%",
            "Alto",
        ],
    ]
    for ex in examples:
        row = _add_example_row(ws, row, ex)

    s = _styles()
    for r in range(row, row + 10):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = s["input_fill"]


def _build_decide(wb: Any, *, lang: str) -> None:
    ws = wb.create_sheet(SHEET_NAMES[4])

    instruction = (
        "DECIDIR: Com base nos cenarios, qual a sua decisao? "
        "Registre o tipo (alerta, recomendacao, acao), sua confianca, "
        "e o custo de errar (falso positivo e falso negativo)."
        if lang == "pt"
        else "DECIDE: Based on scenarios, what is your decision? "
             "Record the type, confidence, and cost of error."
    )

    headers = ["Decisao", "Tipo", "Recomendacao", "Confianca %", "Risco",
               "Custo Falso Positivo", "Custo Falso Negativo"]
    widths = [COL_WIDTH_WIDE, COL_WIDTH_NARROW, COL_WIDTH_WIDE, COL_WIDTH_NARROW,
              COL_WIDTH_NARROW, COL_WIDTH_WIDE, COL_WIDTH_WIDE]

    _set_col_widths(ws, widths)
    row = _add_instruction(ws, 1, instruction, len(headers))
    row = _add_headers(ws, row, headers)

    examples = [
        [
            "Antecipar compra de cafe para Dez",
            "recomendacao",
            "Comprar 50kg agora a 32.77/kg ao inves de esperar possivel alta",
            "65%",
            "Medio",
            "Paguei mais caro se preco cair (perda ~R$240)",
            "Se nao comprar e preco subir, pago ~R$160 a mais depois",
        ],
    ]
    for ex in examples:
        row = _add_example_row(ws, row, ex)

    s = _styles()
    for r in range(row, row + 10):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = s["input_fill"]


def _build_feedback(wb: Any, *, lang: str) -> None:
    ws = wb.create_sheet(SHEET_NAMES[5])

    instruction = (
        "FEEDBACK: Depois que o tempo passou, registre o que realmente aconteceu. "
        "Compare com o que voce previa. Calcule o erro. Anote o que aprendeu."
        if lang == "pt"
        else "FEEDBACK: After the fact, record what actually happened. "
             "Compare with your prediction. Calculate error. Note lessons learned."
    )

    headers = ["Decisao Original", "Metrica", "Previsto", "Real",
               "Erro %", "Acertou?", "Licao Aprendida"]
    widths = [COL_WIDTH_WIDE, COL_WIDTH_MEDIUM, COL_WIDTH_NARROW, COL_WIDTH_NARROW,
              COL_WIDTH_NARROW, COL_WIDTH_NARROW, COL_WIDTH_WIDE]

    _set_col_widths(ws, widths)
    row = _add_instruction(ws, 1, instruction, len(headers))
    row = _add_headers(ws, row, headers)

    examples = [
        [
            "Antecipar compra de cafe",
            "Preco cafe/kg",
            32.77,
            35.20,
            "7.4%",
            "Sim",
            "A seca realmente impactou. Decisao de antecipar economizou ~R$120.",
        ],
    ]
    for ex in examples:
        row = _add_example_row(ws, row, ex)

    s = _styles()
    for r in range(row, row + 10):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = s["input_fill"]


def _build_dashboard(wb: Any, *, lang: str) -> None:
    ws = wb.create_sheet(SHEET_NAMES[6])

    instruction = (
        "DASHBOARD: Visao consolidada. Quantas decisoes voce tomou? "
        "Quantas acertou? Qual seu indice de acerto? "
        "Preencha conforme completa ciclos nas abas anteriores."
        if lang == "pt"
        else "DASHBOARD: Consolidated view. How many decisions? How many correct? "
             "Fill as you complete cycles in previous tabs."
    )

    headers = ["Metrica", "Valor"]
    widths = [COL_WIDTH_WIDE, COL_WIDTH_MEDIUM]

    _set_col_widths(ws, widths)
    row = _add_instruction(ws, 1, instruction, len(headers))
    row = _add_headers(ws, row, headers)

    metrics = [
        ("Total de decisoes", 1),
        ("Decisoes corretas", 1),
        ("Indice de acerto", "100%"),
        ("Erro medio", "7.4%"),
        ("Melhor decisao", "Antecipacao de compra cafe"),
        ("Maior erro", "—"),
        ("Ciclos completos", 1),
    ]
    s = _styles()
    for label, val in metrics:
        cell_label = ws.cell(row=row, column=1, value=label)
        cell_label.font = s["bold"]
        cell_val = ws.cell(row=row, column=2, value=val)
        cell_val.fill = s["input_fill"]
        row += 1


def read_sheet_as_json(xlsx_path: Path) -> dict[str, Any]:
    """Read a filled xlsx template and return a dict compatible with ugear contracts."""
    from openpyxl import load_workbook

    wb = load_workbook(str(xlsx_path), data_only=True)

    result: dict[str, Any] = {}

    if SHEET_NAMES[0] in wb.sheetnames:
        result["observations"] = _read_table(wb[SHEET_NAMES[0]])

    if SHEET_NAMES[1] in wb.sheetnames:
        result["compressions"] = _read_table(wb[SHEET_NAMES[1]])

    if SHEET_NAMES[2] in wb.sheetnames:
        result["hypotheses"] = _read_table(wb[SHEET_NAMES[2]])

    if SHEET_NAMES[3] in wb.sheetnames:
        result["scenarios"] = _read_table(wb[SHEET_NAMES[3]])

    if SHEET_NAMES[4] in wb.sheetnames:
        result["decisions"] = _read_table(wb[SHEET_NAMES[4]])

    if SHEET_NAMES[5] in wb.sheetnames:
        result["feedback"] = _read_table(wb[SHEET_NAMES[5]])

    if SHEET_NAMES[6] in wb.sheetnames:
        result["dashboard"] = _read_table(wb[SHEET_NAMES[6]])

    return result


def _read_table(ws: Any) -> list[dict[str, Any]]:
    """Read a sheet as a list of dicts, skipping instruction rows."""
    rows_iter = ws.iter_rows(values_only=True)
    headers: list[str] = []

    for row_values in rows_iter:
        non_empty = [v for v in row_values if v is not None]
        if len(non_empty) >= MIN_HEADER_COLS and all(isinstance(v, str) for v in non_empty):
            candidate = [str(v).strip() for v in row_values if v is not None]
            if len(candidate) >= MIN_HEADER_COLS:
                headers = candidate
                break

    if not headers:
        return []

    records: list[dict[str, Any]] = []
    for row_values in rows_iter:
        values = list(row_values)[:len(headers)]
        if all(v is None for v in values):
            continue
        record = {}
        for i, h in enumerate(headers):
            val = values[i] if i < len(values) else None
            record[h] = val
        if any(v is not None for v in record.values()):
            records.append(record)

    return records
